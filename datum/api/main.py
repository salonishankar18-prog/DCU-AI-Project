"""
api/main.py — FastAPI routes for Datum.

Thin layer. Every measurement comes from core/, every judgement from review/.
This file only moves data between them and the browser.

The Anthropic key is read server-side from .env and is never included in any
response body.
"""

import asyncio
import json
import os
import queue
import shutil
import threading
import uuid
from pathlib import Path

from dotenv import load_dotenv
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from core.batch import find_plans, run_batch, zip_output

load_dotenv()

ROOT = Path(__file__).resolve().parent.parent
WEB = ROOT / "web"
DEFAULT_DATASET = Path(os.getenv("DATASET_ROOT", ROOT / "data" / "raw"))
OUT_ROOT = Path(os.getenv("OUTPUT_ROOT", ROOT / "data" / "out"))

app = FastAPI(title="Datum", version="1.0")

# In-process job registry for the SSE progress stream. One batch at a time is
# all the UI can drive, but keying by id keeps a stale browser tab from
# attaching to a newer run.
JOBS: dict[str, dict] = {}


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _manifest_path() -> Path:
    return OUT_ROOT / "manifest.json"


def read_manifest() -> list[dict]:
    p = _manifest_path()
    if not p.exists():
        return []
    try:
        with open(p) as fh:
            return json.load(fh)
    except (json.JSONDecodeError, OSError):
        return []


def get_record(plan_id: str) -> dict:
    """The manifest record for one plan, or 404."""
    for rec in read_manifest():
        if rec.get("plan_id") == plan_id:
            return rec
    raise HTTPException(404, f"plan '{plan_id}' not found — run the batch first")


def load_full_record(plan_id: str) -> dict:
    """Per-plan dimensions.json, which keeps the room polygons the manifest drops.

    The geometry checks need the polygons; the manifest is deliberately slim.
    """
    p = OUT_ROOT / plan_id / "dimensions.json"
    if not p.exists():
        raise HTTPException(404, f"no dimensions.json for '{plan_id}'")
    with open(p) as fh:
        return json.load(fh)


def _resolve_root(root: str | None) -> Path:
    path = Path(root).expanduser() if root else DEFAULT_DATASET
    if not path.exists():
        raise HTTPException(400, f"folder not found: {path}")
    if not path.is_dir():
        raise HTTPException(400, f"not a folder: {path}")
    return path


# ---------------------------------------------------------------------------
# folder upload
# ---------------------------------------------------------------------------

@app.post("/api/upload")
async def api_upload(files: list[UploadFile] = File(...)):
    """Stage a folder picked on the user's own computer as a scan root.

    A browser can't hand the server a real filesystem path for an arbitrary
    local folder, so the UI uploads its files instead (an <input webkitdirectory>
    preserves each file's relative path) and they land here, under
    data/uploads/<id>/. `find_plans` walks the whole tree, so it doesn't matter
    how deep the plan folders sit inside the uploaded structure.
    """
    if not files:
        raise HTTPException(400, "no files received")

    dest_root = (ROOT / "data" / "uploads" / uuid.uuid4().hex[:10]).resolve()
    saved = 0
    for f in files:
        parts = [p for p in Path(f.filename or "").as_posix().split("/")
                 if p not in ("", ".", "..")]
        if not parts:
            continue
        dest = dest_root.joinpath(*parts)
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as out:
            shutil.copyfileobj(f.file, out)
        saved += 1
        await f.close()

    if not saved:
        raise HTTPException(400, "no valid files in that folder")

    return {"root": str(dest_root), "files": saved}


# ---------------------------------------------------------------------------
# Tab 1 — extract dimensions
# ---------------------------------------------------------------------------

@app.post("/api/scan")
async def api_scan(payload: dict):
    """Walk a dataset folder and report what is there. Nothing is written."""
    root = _resolve_root(payload.get("root"))
    plans = find_plans(root)
    png_count = sum(len(p["pngs"]) for p in plans)
    return {
        "root": str(root),
        "plans": len(plans),
        "pngs": png_count,
        "plan_ids": [p["plan_id"] for p in plans],
        "already_annotated": len(read_manifest()),
    }


@app.post("/api/annotate")
async def api_annotate(payload: dict):
    """Start a batch run. Returns a job id; progress arrives on the SSE stream.

    The batch itself is synchronous and CPU-bound, so it runs on a worker thread
    and pushes progress events into a queue the stream drains.
    """
    root = _resolve_root(payload.get("root"))
    units = payload.get("units", "m")
    style = payload.get("style", "dimension_lines")
    room_labels = payload.get("room_labels", True)
    limit = payload.get("limit") or None
    resume = bool(payload.get("resume", True))

    if not resume:
        # A fresh run must not resume from a manifest written under different
        # settings — otherwise "Annotate all" silently does nothing.
        mp = _manifest_path()
        if mp.exists():
            mp.unlink()

    job_id = uuid.uuid4().hex[:12]
    q: queue.Queue = queue.Queue()
    JOBS[job_id] = {"queue": q, "done": False, "summary": None}

    def progress(done, total, plan_id, status):
        q.put({"type": "progress", "done": done, "total": total,
               "plan_id": plan_id, "status": status})

    def worker():
        try:
            summary = run_batch(
                str(root), str(OUT_ROOT), units=units, style=style,
                room_labels=room_labels, limit=limit, resume=resume,
                progress=progress,
            )
            JOBS[job_id]["summary"] = summary
            q.put({"type": "done", **summary})
        except Exception as exc:                                    # noqa: BLE001
            q.put({"type": "error", "error": f"{type(exc).__name__}: {exc}"})
        finally:
            JOBS[job_id]["done"] = True
            q.put(None)                                             # stream sentinel

    threading.Thread(target=worker, daemon=True).start()

    total = len(find_plans(root)) if not limit else min(limit, len(find_plans(root)))
    return {"job_id": job_id, "total": total}


@app.get("/api/jobs/{job_id}/events")
async def api_job_events(job_id: str, request: Request):
    """Server-sent events for one background job. Real progress, not a fake timer."""
    job = JOBS.get(job_id)
    if job is None:
        raise HTTPException(404, "unknown job")

    async def stream():
        q: queue.Queue = job["queue"]
        while True:
            if await request.is_disconnected():
                break
            try:
                item = q.get_nowait()
            except queue.Empty:
                await asyncio.sleep(0.05)
                continue
            if item is None:
                break
            yield f"data: {json.dumps(item)}\n\n"
        JOBS.pop(job_id, None)

    return StreamingResponse(stream(), media_type="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
    })


@app.get("/api/plans")
async def api_plans():
    """The manifest — the interface between Part 1 and Part 2."""
    manifest = read_manifest()
    counts: dict[str, int] = {}
    for rec in manifest:
        counts[rec.get("status", "?")] = counts.get(rec.get("status", "?"), 0) + 1
    return {"plans": manifest, "counts": counts, "out_root": str(OUT_ROOT)}


@app.get("/api/plans/{plan_id}/png")
async def api_plan_png(plan_id: str, file: str | None = None):
    """Serve an annotated PNG. Defaults to the first output of the plan."""
    rec = get_record(plan_id)
    outputs = rec.get("outputs") or []
    if not outputs:
        raise HTTPException(404, f"'{plan_id}' has no annotated output")

    name = file or outputs[0]["output_png"]
    if "/" in name or "\\" in name or ".." in name:
        raise HTTPException(400, "bad filename")

    path = OUT_ROOT / plan_id / name
    if not path.exists():
        raise HTTPException(404, f"{name} not found")
    return FileResponse(path, media_type="image/png")


@app.get("/api/export/csv")
async def api_export_csv():
    path = OUT_ROOT / "dimensions.csv"
    if not path.exists():
        raise HTTPException(404, "no CSV yet — run the batch first")
    return FileResponse(path, media_type="text/csv", filename="datum_dimensions.csv")


@app.get("/api/export/zip")
async def api_export_zip():
    """Originals + annotated copies + per-plan dimensions.json, in one file."""
    if not _manifest_path().exists():
        raise HTTPException(404, "nothing to export yet — run the batch first")
    zip_path = ROOT / "data" / "datum_export.zip"
    zip_output(str(OUT_ROOT), str(zip_path))
    return FileResponse(zip_path, media_type="application/zip", filename="datum_export.zip")


# ---------------------------------------------------------------------------
# Tab 2 — access review
# ---------------------------------------------------------------------------

def _svg_for(plan_id: str) -> str:
    """The copy of model.svg written into the output folder. data/raw is never touched."""
    p = OUT_ROOT / plan_id / "model.svg"
    if not p.exists():
        raise HTTPException(404, f"no model.svg for '{plan_id}' — run the batch first")
    return str(p)


@app.get("/api/tgdm/status")
async def api_tgdm_status():
    from review.agent import MODEL
    from review.tgdm_index import DEFAULT_PDF, get_index
    try:
        idx = get_index()
        return {
            "indexed": len(idx) > 0,
            "clauses": len(idx),
            "pdf": DEFAULT_PDF.name,
            "model": MODEL,
            "key_configured": bool(os.getenv("ANTHROPIC_API_KEY", "").strip()
                                   and not os.getenv("ANTHROPIC_API_KEY", "").startswith("sk-ant-your-key")),
        }
    except Exception as exc:                                        # noqa: BLE001
        return {"indexed": False, "clauses": 0, "error": str(exc), "model": MODEL}


@app.get("/api/plans/{plan_id}/checks")
async def api_checks(plan_id: str):
    """Deterministic geometry — Python's numbers, before Claude sees them."""
    from review.rules import plan_geometry
    get_record(plan_id)
    try:
        return plan_geometry(_svg_for(plan_id), plan_id)
    except HTTPException:
        raise
    except Exception as exc:                                        # noqa: BLE001
        raise HTTPException(500, f"geometry failed: {type(exc).__name__}: {exc}")


@app.post("/api/review")
async def api_review(payload: dict):
    """{plan_id, question} -> Claude's answer with real clause citations."""
    from review.agent import NotConfigured, answer_question

    plan_id = (payload.get("plan_id") or "").strip()
    question = (payload.get("question") or "").strip()
    if not plan_id or not question:
        raise HTTPException(400, "plan_id and question are required")
    get_record(plan_id)

    try:
        return await asyncio.to_thread(
            answer_question, _svg_for(plan_id), plan_id, question,
            payload.get("history") or [],
        )
    except NotConfigured as exc:
        raise HTTPException(503, str(exc))
    except HTTPException:
        raise
    except Exception as exc:                                        # noqa: BLE001
        raise HTTPException(500, f"{type(exc).__name__}: {exc}")


@app.get("/api/plans/{plan_id}/verdict")
async def api_cached_verdict(plan_id: str):
    """The stored determination, if this plan has already been run. Never calls the API."""
    rec = get_record(plan_id)
    if not rec.get("verdict"):
        raise HTTPException(404, "no determination stored for this plan")
    cached = dict(rec["verdict"])
    cached["cached"] = True
    return cached


@app.post("/api/verdict")
async def api_verdict(payload: dict):
    """{plan_id} -> structured determination. Cached in the manifest after first run."""
    from review.agent import NotConfigured, verdict

    plan_id = (payload.get("plan_id") or "").strip()
    if not plan_id:
        raise HTTPException(400, "plan_id is required")
    rec = get_record(plan_id)

    # The demo replays from cache rather than re-billing every click.
    if not payload.get("refresh") and rec.get("verdict"):
        cached = dict(rec["verdict"])
        cached["cached"] = True
        return cached

    try:
        result = await asyncio.to_thread(verdict, _svg_for(plan_id), plan_id)
    except NotConfigured as exc:
        raise HTTPException(503, str(exc))
    except HTTPException:
        raise
    except Exception as exc:                                        # noqa: BLE001
        raise HTTPException(500, f"{type(exc).__name__}: {exc}")

    _cache_verdict(plan_id, result)
    result["cached"] = False
    return result


# A "Run all determinations" click bills real API usage per plan. Capped so a
# large dataset can never run away with the configured key's spend — once
# cumulative cost crosses this, plans still queued are skipped rather than run.
VERDICT_ALL_BUDGET_USD = 2.00


@app.post("/api/verdict/all")
async def api_verdict_all(payload: dict):
    """Run the determination for every plan in the folder, streaming progress.

    Plans already determined are skipped unless `refresh` is set, so an
    interrupted run resumes instead of re-billing what it already paid for.
    A small worker pool keeps a 100-plan run to minutes without hammering the
    rate limit. Cumulative spend is capped at VERDICT_ALL_BUDGET_USD — once
    that's crossed, remaining plans are marked skipped_budget rather than run.
    """
    from review.agent import NotConfigured, verdict

    refresh = bool(payload.get("refresh"))
    limit = payload.get("limit") or None
    workers = max(1, min(int(payload.get("workers") or 3), 6))

    manifest = read_manifest()
    todo = [r["plan_id"] for r in manifest if r.get("status") != "failed"]
    if not refresh:
        done = {r["plan_id"] for r in manifest if r.get("verdict")}
        todo = [p for p in todo if p not in done]
    if limit:
        todo = todo[:limit]

    job_id = uuid.uuid4().hex[:12]
    q: queue.Queue = queue.Queue()
    JOBS[job_id] = {"queue": q, "done": False, "summary": None}

    def worker():
        from concurrent.futures import ThreadPoolExecutor, as_completed
        counts: dict[str, int] = {}
        total = len(todo)
        lock = threading.Lock()
        done_n = 0
        spent = {"total": 0.0}
        budget_hit = threading.Event()

        def one(pid):
            if budget_hit.is_set():
                return pid, None, None, "skipped_budget"
            try:
                result = verdict(_svg_for(pid), pid)
            except NotConfigured:
                raise
            except Exception as exc:                               # noqa: BLE001
                return pid, None, f"{type(exc).__name__}: {exc}", None
            cost = ((result or {}).get("usage") or {}).get("cost_usd") or 0.0
            with lock:
                spent["total"] += cost
                if spent["total"] >= VERDICT_ALL_BUDGET_USD:
                    budget_hit.set()
            return pid, result, None, None

        try:
            if not total:
                q.put({"type": "done", "total": 0, "counts": {},
                       "spent_usd": 0.0, "budget_cap_usd": VERDICT_ALL_BUDGET_USD,
                       "message": "every plan already has a determination"})
                return
            with ThreadPoolExecutor(max_workers=workers) as pool:
                futures = [pool.submit(one, p) for p in todo]
                for fut in as_completed(futures):
                    pid, result, err, skipped = fut.result()
                    with lock:
                        done_n += 1
                        n = done_n
                    if skipped == "skipped_budget":
                        key = "skipped_budget"
                    elif result is not None:
                        _cache_verdict(pid, result)
                        key = result.get("determination", "?")
                    else:
                        key = "error"
                    counts[key] = counts.get(key, 0) + 1
                    q.put({"type": "progress", "done": n, "total": total,
                           "plan_id": pid, "status": key, "error": err})
            q.put({"type": "done", "total": total, "counts": counts,
                   "spent_usd": round(spent["total"], 4),
                   "budget_cap_usd": VERDICT_ALL_BUDGET_USD})
        except NotConfigured as exc:
            q.put({"type": "error", "error": str(exc)})
        except Exception as exc:                                    # noqa: BLE001
            q.put({"type": "error", "error": f"{type(exc).__name__}: {exc}"})
        finally:
            JOBS[job_id]["done"] = True
            q.put(None)

    threading.Thread(target=worker, daemon=True).start()
    return {"job_id": job_id, "total": len(todo)}


# The batch runner writes determinations from several threads at once, and each
# write is a read-modify-write of the whole manifest. Without this lock a
# concurrent pair silently loses one of the two verdicts.
_MANIFEST_LOCK = threading.Lock()


def _cache_verdict(plan_id: str, result: dict) -> None:
    """Write the determination back into the manifest, per the rate-limit plan."""
    with _MANIFEST_LOCK:
        path = _manifest_path()
        manifest = read_manifest()
        if not manifest:
            return
        for rec in manifest:
            if rec.get("plan_id") == plan_id:
                rec["verdict"] = result
                break
        tmp = path.with_suffix(".json.tmp")
        with open(tmp, "w") as fh:
            json.dump(manifest, fh, indent=2)
        tmp.replace(path)


# ---------------------------------------------------------------------------
# Excel export — everything, in one workbook
# ---------------------------------------------------------------------------

def _geometry_for(plan_id: str) -> dict:
    """Per-plan geometry, cached beside the annotated output."""
    cache = OUT_ROOT / plan_id / "geometry.json"
    if cache.exists():
        try:
            with open(cache) as fh:
                return json.load(fh)
        except (json.JSONDecodeError, OSError):
            pass
    from review.rules import plan_geometry
    geo = plan_geometry(_svg_for(plan_id), plan_id)
    try:
        with open(cache, "w") as fh:
            json.dump(geo, fh, indent=2)
    except OSError:
        pass
    return geo


def _build_workbook(path: Path) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="14202A")      # structural ink
    dim_font = Font(color="E0532B", bold=True)              # measurement ink

    def sheet(title, headers, rows, measure_cols=()):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = head_font, head_fill
            cell.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(r)
        # The colour rule holds in the spreadsheet too: orange marks a value
        # describing a real-world distance or area, and nothing else.
        for col in measure_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col).font = dim_font
        for c in range(1, len(headers) + 1):
            longest = max([len(str(headers[c - 1]))] +
                          [len(str(row[c - 1])) for row in rows[:400]
                           if c - 1 < len(row) and row[c - 1] is not None] or [10])
            ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 2, 10), 60)
        ws.freeze_panes = "A2"
        return ws

    manifest = read_manifest()

    plans, rooms, doors, corridors, checks, clearances = [], [], [], [], [], []
    for rec in manifest:
        pid = rec.get("plan_id")
        v = rec.get("verdict") or {}
        vchecks = v.get("checks") or []
        plans.append([
            pid, rec.get("status"), rec.get("scale_source"), rec.get("px_per_m"),
            rec.get("overall_width_m"), rec.get("overall_depth_m"),
            rec.get("internal_area_m2"), len(rec.get("rooms") or []),
            v.get("determination", "not run"),
            sum(1 for c in vchecks if c.get("result") == "pass"),
            sum(1 for c in vchecks if c.get("result") == "fail"),
            sum(1 for c in vchecks if c.get("result") == "undetermined"),
            v.get("summary", ""),
        ])
        for room in rec.get("rooms") or []:
            rooms.append([pid, room.get("code"), room.get("name"), room.get("category"),
                          room.get("width_m"), room.get("depth_m"), room.get("area_m2"),
                          room.get("is_rectangular"), room.get("outdoor")])
        for c in vchecks:
            checks.append([pid, v.get("determination"), c.get("item"),
                           c.get("measured_display"), c.get("result"),
                           c.get("clause"), c.get("note")])

        if rec.get("status") == "failed":
            continue
        try:
            geo = _geometry_for(pid)
        except Exception:                                           # noqa: BLE001
            continue
        for d in geo.get("doors", []):
            doors.append([pid, d.get("id"), " → ".join(d.get("between") or []),
                          d.get("clear_width_mm"), d.get("wall_thickness_mm"),
                          d.get("external"), d.get("method")])
        for co in geo.get("corridors", []):
            corridors.append([pid, co.get("id"), co.get("code"), co.get("name"),
                              co.get("min_clear_width_mm"),
                              co.get("min_clear_width_less_furniture_mm"),
                              co.get("area_m2"), co.get("identified_by"),
                              co.get("probable_only")])
        for r in geo.get("rooms", []):
            clearances.append([pid, r.get("code"), r.get("name"), r.get("area_m2"),
                               r.get("free_circle_mm"),
                               r.get("free_circle_ignoring_furniture_mm"),
                               r.get("min_width_mm"), r.get("furniture_present"),
                               r.get("outdoor")])

    sheet("Plans",
          ["Plan ID", "Status", "Scale source", "Scale px/m", "Width m", "Depth m",
           "Floor area m2", "Rooms", "Determination", "Pass", "Fail", "Undetermined",
           "Summary"],
          plans, measure_cols=(5, 6, 7))
    sheet("Rooms",
          ["Plan ID", "Code", "Room", "Category", "Width m", "Depth m", "Area m2",
           "Rectangular", "Outdoor"],
          rooms, measure_cols=(5, 6, 7))
    sheet("Doors",
          ["Plan ID", "Door", "Between", "Clear opening mm", "Wall thickness mm",
           "External", "Method"],
          doors, measure_cols=(4, 5))
    sheet("Corridors",
          ["Plan ID", "ID", "Code", "Name", "Min clear width mm",
           "Min width less furniture mm", "Area m2", "Identified by", "Heuristic only"],
          corridors, measure_cols=(5, 6, 7))
    sheet("Room clearances",
          ["Plan ID", "Code", "Room", "Area m2", "Free circle mm",
           "Free circle ignoring furniture mm", "Min width mm", "Furniture present",
           "Outdoor"],
          clearances, measure_cols=(4, 5, 6, 7))
    sheet("Access checks",
          ["Plan ID", "Determination", "Check", "Measured", "Result",
           "TGD M clause", "Note"],
          checks, measure_cols=(4,))

    notes = wb.create_sheet("Method", 0)
    for line in [
        ["Datum — floor plan dimension & access review"],
        [],
        ["Dimensions", "Read from the hidden DimensionMeasureLabel in each room of model.svg,"],
        ["", "cross-checked against the drawn polygon. Area is true polygon area."],
        ["Door clear opening", "Span of the SVG Threshold rectangle parallel to the wall face."],
        ["", "This is a STRUCTURAL clear opening between wall faces. TGD M measures"],
        ["", "effective clear width past the open leaf, which is somewhat narrower."],
        ["Corridor width", "Narrowest cross-section perpendicular to the space's long axis."],
        ["Free circle", "Largest circle fitting inside the room with fixed furniture subtracted."],
        [],
        ["Determinations", "Python computes every measurement; Claude judges it against verbatim"],
        ["", "clause text retrieved from TGD M 2022. No threshold is hard-coded."],
        ["", "Three outcomes: pass / fail / undetermined. Undetermined means the plan"],
        ["", "view cannot show it — the note names the drawing that would."],
        ["Excluded", "Plans whose scale fell through to the convention fallback are excluded"],
        ["", "from determination; their measurements are not reliable."],
        [],
        ["Orange", "marks a value describing a real-world distance or area, and nothing else."],
    ]:
        notes.append(line)
    notes["A1"].font = Font(bold=True, size=13)
    notes.column_dimensions["A"].width = 20
    notes.column_dimensions["B"].width = 95

    del wb["Sheet"]
    wb.save(path)
    return path


@app.get("/api/export/xlsx")
async def api_export_xlsx():
    """One workbook: plans, rooms, doors, corridors, clearances, access checks."""
    if not _manifest_path().exists():
        raise HTTPException(404, "nothing to export yet — run the batch first")
    path = ROOT / "data" / "datum_report.xlsx"
    await asyncio.to_thread(_build_workbook, path)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="datum_report.xlsx",
    )


# ---------------------------------------------------------------------------
# static UI
# ---------------------------------------------------------------------------

@app.get("/")
async def index():
    return FileResponse(WEB / "index.html")


@app.get("/api/health")
async def health():
    return {
        "ok": True,
        "dataset_default": str(DEFAULT_DATASET),
        "out_root": str(OUT_ROOT),
        "annotated_plans": len(read_manifest()),
    }


app.mount("/web", StaticFiles(directory=str(WEB)), name="web")
